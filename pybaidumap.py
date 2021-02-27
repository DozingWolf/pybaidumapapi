import requests
import os
import json
import openpyxl
from configparser import ConfigParser

# 一级分类：房地产
# 二级分类：写字楼、住宅区、宿舍、内部楼栋、其他
# 121.393658,31.220479
# api demo
# http://api.map.baidu.com/place/v2/search?query=%E4%BD%8F%E5%AE%85%E5%8C%BA&location=31.222042,121.390664&radius=700&output=json&page_size=20&page_num=0&ak=
# http://api.map.baidu.com/geocoding/v3/?address=%E5%8C%97%E6%96%B0%E6%B3%BE&city=%E4%B8%8A%E6%B5%B7%E5%B8%82&output=json&ak=
# https://sh.lianjia.com/ershoufang/?sug=%E7%95%85%E5%9B%AD

paraPath = './conf/para.conf'
paraLoader = ConfigParser()
paraLoader.read(paraPath)
ak_key = paraLoader.get('baidu_map','ak')
output_path = paraLoader.get('output','path')

class excelOper(object):
    def __init__(self,path,fname):
        self.__filepath = path+fname
        self.__workbook = openpyxl.Workbook()
    def activeWorkSheet(self,sname):
        self.__workbook.create_sheet(sname,0)
        self.__activeSheet = self.__workbook.active
        self.__activeSheet.title = sname
    def insertIntoWorkSheet(self,value):
        for i,row in enumerate(value):
            for j,col in enumerate(value[i]):
                self.__activeSheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    def saveWorkBook(self):
        self.__workbook.save(self.__filepath)
        

def write_excel_xlsx(path, sheet_name, value):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i,row in enumerate(value):
        for j,col in enumerate(value[i]):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")

def getLatitudeandLongitude(ak_key,pname,city,rctype='bd09ll'):
    resultMessage = ''
    getLatitudeandLongitudeurl = 'http://api.map.baidu.com/geocoding/v3/'
    urlHeaders = {
        'Connection':'keep-alive',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.182 Safari/537.36 Edg/88.0.705.74',
        'Accept-Encoding':'gzip, deflate'
    }
    getLatandLonPara = {
        'address':pname,
        'city':city,
        'ret_coordtype':rctype,
        'output':'json',
        'ak':ak_key
    }

    try:
        r = requests.get(getLatitudeandLongitudeurl,params=getLatandLonPara,headers=urlHeaders)
        decodejson = json.loads(r.text)
        if decodejson.get('result') == []:
            raise Exception('getLatitudeandLongitude_API return No Data Found!')
        else:
            resultMessage = ','.join([str(decodejson.get('result').get('location').get('lat')),str(decodejson.get('result').get('location').get('lng'))])
    except Exception as err:
        print(err)
    
    return resultMessage
    
    

def getPOI(ak_key,r=700,l='0,0',pagesize=20,querydata=u'住宅区'):
    resultMessage = []
    cycleFlag = 0
    page=0
    getPOIurl = 'http://api.map.baidu.com/place/v2/search'
    urlHeaders = {
        'Connection':'keep-alive',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.182 Safari/537.36 Edg/88.0.705.74',
        'Accept-Encoding':'gzip, deflate'
    }
    getPOIurlPara = {
        'query':querydata,
        'radius':r,
        'page_size':pagesize,
        'page_num':page,
        'location':l,
        'ak':ak_key,
        'output':'json'
    }
    while cycleFlag == 0 :
        try:
            r = requests.get(getPOIurl,params=getPOIurlPara,headers=urlHeaders)
            decodejson = json.loads(r.text)
            if decodejson.get('results') != []:
                resultMessage.extend(decodejson.get('results'))
                newPageNo = int(getPOIurlPara.get('page_num'))+1
                getPOIurlPara.update({'page_num':newPageNo})
                # print(decodejson.get('results'))
            else:
                print('getPOI return No data found!')
                cycleFlag = 1
        except Exception as err:
            cycleFlag = 1
            print(err)
    return resultMessage

def dataTrans(dataclassification,value):
    if dataclassification == 'POI' :
        formatResult = []
        for k,v in enumerate(value):
            temp_formatResult = []
            temp_formatResult.extend([v.get('name')])
            temp_formatResult.extend([v.get('address')])
            temp_formatResult.extend([v.get('location').get('lat')])
            temp_formatResult.extend([v.get('location').get('lng')])
            formatResult.append(temp_formatResult)
        return formatResult

def __main__():
    apiResult = getPOI(ak_key=ak_key,l='31.222042,121.390664')
    formatResult = []

    for k,v in enumerate(apiResult):
        temp_formatResult = []
        temp_formatResult.extend([v.get('name')])
        temp_formatResult.extend([v.get('address')])
        temp_formatResult.extend([v.get('location').get('lat')])
        temp_formatResult.extend([v.get('location').get('lng')])
        formatResult.append(temp_formatResult)

    # write_excel_xlsx(path=output_path,sheet_name='威宁路700M',value=formatResult)

def testFunc():
    dataresult = getPOI(ak_key=ak_key,r=700,l=getLatitudeandLongitude(ak_key=ak_key,pname='威宁路地铁站',city='上海市'))
    activeWorkBook = excelOper(path=output_path,fname='地铁站700m范围楼盘.xlsx')

    formatResult = []
    for k,v in enumerate(dataresult):
        temp_formatResult = []
        temp_formatResult.extend([v.get('name')])
        temp_formatResult.extend([v.get('address')])
        temp_formatResult.extend([v.get('location').get('lat')])
        temp_formatResult.extend([v.get('location').get('lng')])
        formatResult.append(temp_formatResult)
    activeWorkBook.activeWorkSheet(sname='威宁路')
    activeWorkBook.insertIntoWorkSheet(value=formatResult)
    dataresult2 = getPOI(ak_key=ak_key,r=700,l=getLatitudeandLongitude(ak_key=ak_key,pname='北新泾地铁站',city='上海市'))
    formatResult2 = []
    for k,v in enumerate(dataresult2):
        temp_formatResult = []
        temp_formatResult.extend([v.get('name')])
        temp_formatResult.extend([v.get('address')])
        temp_formatResult.extend([v.get('location').get('lat')])
        temp_formatResult.extend([v.get('location').get('lng')])
        formatResult2.append(temp_formatResult)
    activeWorkBook.activeWorkSheet(sname='北新泾')
    activeWorkBook.insertIntoWorkSheet(value=formatResult2)
    activeWorkBook.saveWorkBook()

testFunc()

# poi=getLatitudeandLongitude(ak_key=ak_key,pname='威宁路地铁站',city='上海市')
# dataresult = getPOI(ak_key=ak_key,r=700,l=poi)
# print(dataresult)